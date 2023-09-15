from codigo.app_chilquinta import Scraper_Chilquinta
from openpyxl import load_workbook

def send_notification():
    # Código para enviar correo electrónico de notificación
    print('')

if __name__ == '__main__':
    
    print('Obteniendo credenciales...')
    print('----------------------------------------------------------------------')
        
    credencials = 'C:\\roda\\chilquinta-portal\\config\\credenciales.xlsx'
    libro_accesos = load_workbook(credencials)
    hoja_credenciales = libro_accesos['Hoja1']
        
    for j in hoja_credenciales.iter_rows(2):
        try:
            rut = j[0].value
            passw = j[1].value
            web = j[2].value
            break
        except:
            ('no hay credenciales')
            
    email = rut
    password = passw
    url = web
    driver_path = 'chromedriver.exe'
    
    scraper = Scraper_Chilquinta(url, email, password, driver_path)

    #Primer ingreso, año actual
    print('ingresamos en la clase Scraper_Chilquinta...')
    print('----------------------------------------------------------------------')
    
    # scraper.login()
    # print('hacemos login en el portal...')
    # print('----------------------------------------------------------------------')
    
    # scraper.scrapping_chilquita()
    # print('hacemos scrapping al portal...')
    # print('----------------------------------------------------------------------')
    
    scraper.archivos()
    print('procesamos archivos datos hacia Planilla Formato...')
    print('----------------------------------------------------------------------')
    
    scraper.close()
    print('cerramos el bot...')
    print('----------------------------------------------------------------------')

