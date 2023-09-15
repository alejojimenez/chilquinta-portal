import os
import time
import shutil
import rpa as r
import requests
import glob
import fitz
import re

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.alert import Alert
from openpyxl import load_workbook
from datetime import datetime,timedelta
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

class Scraper_Chilquinta():

    def __init__(self,url, email, password, driver_path):
        print(url, email, password, driver_path)
        self.url = url
        self.email = email
        self.password = password
        self.driver_path = driver_path

    def wait(self, seconds):
        return WebDriverWait(self.driver, seconds)

    def close(self):
        self.driver.close()
        self.driver = None

    def quit(self):
        self.driver.quit()
        self.driver = None

    def dic_datos(self,mes_texto):
        dic = {'ENE':'01','FEB':'02','MAR':'03','ABR':'04','MAY':'05','JUN':'06','JUL':'07',
               'AGO':'08','SEP':'09','OCT':'10','NOV':'11','DIC':'12'}

        mes_mayu = mes_texto.upper()
        mes_oficial = dic[mes_mayu]
        
        return mes_oficial

    def diccionario(self):
        dic_mes = {'01':'Enero','02':'Febrero',
            '03':'Marzo','04':'Abril',
            '05':'Mayo','06':'Junio',
            '07':'Julio','08':'Agosto',
            '09':'Septiembre','10':'Octubre',
            '11':'Noviembre','12':'Diciembre',
            }
        return dic_mes
        
    def login(self):
        
        #driver_exe = 'C:\\roda\\saesa_portal\\domain\\chromedriver.exe'
        credencials = 'C:\\roda\\chilquinta-portal\\config\\credenciales.xlsx'
        driver_exe = './chromedriver.exe'

        print('Entrando en la funcion login...')
        print('----------------------------------------------------------------------')
        
        #Seteo variables
        email = self.email
        url = self.url
        driver_path = self.driver_path
        password  = self.password
        
        #Version 4.9.1 de selenium
        options = webdriver.ChromeOptions()    
        self.driver = webdriver.Chrome(self.driver_path,options=options)

        self.driver.get(self.url)
        self.driver.maximize_window()
        self.driver.implicitly_wait(40)

        # Controlar evento alert() Notificacion
        try:
            alert = WebDriverWait(self.driver, 35).until(EC.alert_is_present())
            alert = Alert(self.driver)
            alert.accept()
                    
        except:
            print("No se encontró ninguna alerta.")  

        #Boton donde se encuentran todos los servicios
        # intentos = 0
        # servicios = True
        # while (servicios):
        #         try:
        #             print('Try en la funcion servicios...', intentos)
        #             print('----------------------------------------------------------------------')
        #             intentos += 1
        #             boton_servicios = WebDriverWait(self.driver, 20).until(
        #             EC.element_to_be_clickable((By.CSS_SELECTOR, 'button.btn-services span.span-services')))
        #             boton_servicios.click()
        #             servicios = False
        #         except:    
        #             print('Exception en la funcion servicios...')
        #             print('----------------------------------------------------------------------')
        #             servicios = intentos <= 3  
            
        #Boton de historial de boletas
        # intentos = 0
        # historial = True
        # while (historial):
        #         try:
        #             print('Try en la funcion historial ...', intentos)
        #             print('----------------------------------------------------------------------')
        #             intentos += 1
        #             boton_menu= self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div[2]/section/section/div[3]/div/div/div/div/div/div[2]/div[2]/div/div[1]/div/a')
        #             time.sleep(2)
        #             boton_menu.click()
        #             historial = False
        #         except:    
        #             print('Exception en la funcion historial ...')
        #             print('----------------------------------------------------------------------')
        #             historial = intentos <= 3
            
        #Botones ID y XPATH que utilizaremos para logear
        selector_username_input = 'input-product'
        selector_password_input = '/html/body/div[1]/div[2]/section/div[2]/div/div/div/div/div/div/div/div/div[2]/div[2]/div[1]/div/div/input'
        selector_ingreso_button = '/html/body/div[1]/div[2]/section/div[2]/div/div/div/div/div/div/div/div/div[2]/button'
                
        #Primero seteamos el usuario
        intentos = 0
        usuario = True
        while (usuario):
                try:
                    print('Try en el ingreso del usuario ...', intentos)
                    print('----------------------------------------------------------------------')
                    intentos += 1
                    element_username = WebDriverWait(self.driver, 20).until(
                        EC.presence_of_element_located((By.ID, selector_username_input)))
                    time.sleep(5)
                    element_username.clear()
                    element_username.click()
                    element_username.send_keys(email)
                    usuario = False
                except:    
                    print('Exception en el ingreso del usuario  ...')
                    print('----------------------------------------------------------------------')
                    usuario = intentos <= 3
            
        #Luego seteamos la password para ingreso
        intentos = 0
        clave = True
        while (clave):
                try:
                    print('Try en el ingreso de la clave ...', intentos)
                    print('----------------------------------------------------------------------')
                    intentos += 1
                    element_username = WebDriverWait(self.driver, 20).until(
                        EC.presence_of_element_located((By.XPATH, selector_password_input)))
                    time.sleep(5)
                    element_username.clear()
                    element_username.click()
                    element_username.send_keys(password)
                    clave = False
                except:    
                    print('Exception en el ingreso de la clave ...')
                    print('----------------------------------------------------------------------')
                    clave = intentos <= 3
        
        intentos = 0
        ingreso = True
        while (ingreso):
                try:
                    print('Try en el ingreso de la clave ...', intentos)
                    print('----------------------------------------------------------------------')
                    intentos += 1
                    boton_ingreso = WebDriverWait(self.driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, selector_ingreso_button)))
                    boton_ingreso.click()
                    ingreso = False
                except:    
                    print('Exception en el ingreso de la clave ...')
                    print('----------------------------------------------------------------------')
                    ingreso = intentos <= 3        
        
        print('Ya logramos ingresar, ahora vamos a buscar las factruras')
        time.sleep(7)
        
    def scrapping_chilquita(self):

        #Buscamos la tabla que contiene las boletas
        intentos = 0
        tabla_total= True
        while (tabla_total):
            try:
                print('Try localizando la tabla con los datos ...', intentos)
                print('----------------------------------------------------------------------')
                intentos += 1
                WebDriverWait(self.driver, 30).until(
                EC.presence_of_element_located((By.XPATH,'//*[@id="table-last-invoice-desktop"]/tbody')))
                self.driver.execute_script("window.scrollBy(0, 350)")
                tabla_total = False
            except:    
                print('Exception en el ingreso de la clave ...')
                print('----------------------------------------------------------------------')
                tabla_total = intentos <= 3    
                                   
        #Determinamos el largo de la tabla, que limitara la cantidad de descargas
        intentos = 0
        largo = True
        while (largo):
            try:
                print('Try localizando la tabla con los datos ...', intentos)
                print('----------------------------------------------------------------------')
                intentos += 1
                tabla= self.driver.find_element(By.XPATH,'//*[@id="table-last-invoice-desktop"]/tbody')
                filas = tabla.find_elements(By.TAG_NAME, "tr")
                cantidad_facturas = len(filas)
                largo = False
            except:    
                print('Exception en el ingreso de la clave ...')
                print('----------------------------------------------------------------------')
                largo = intentos <= 3    
             
        i = 1            
        while i < cantidad_facturas+1:
            
            #Obtenemos el numero de factura desde el texto
            intentos = 0
            n_factura = True
            while (n_factura) and intentos <= 5:
                try:
                    print('Try localizando el numero de factura en la tabla ...', intentos)
                    print('----------------------------------------------------------------------')
                    intentos += 1
                    boleta= self.driver.find_element(By.XPATH,f'/html/body/div[1]/div[2]/section/div[2]/div/div/div[2]/div[2]/div/table/tbody/tr[{i}]/td[1]')
                    valor_boleta = boleta.text
                    print(f'el valor de boleta es {valor_boleta}')
                    n_factura = False
                except:    
                    print('Exception localizando el numero de factura en la tabla...')
                    print('----------------------------------------------------------------------')
                    intentos += 1
                        
            #Obtenemos la fecha y dividimos dicho valor para tener mes y año            
            reintentos = 0            
            fecha_oficial = True
            while (fecha_oficial) and reintentos <= 5:
                try:
                    print('Try localizando fecha en la tabla ...', reintentos)
                    print('----------------------------------------------------------------------')
                    reintentos += 1
                    fecha_completa = self.driver.find_element(By.XPATH,f'/html/body/div[1]/div[2]/section/div[2]/div/div/div[2]/div[2]/div/table/tbody/tr[{i}]/td[3]')
                    fecha_texto = fecha_completa.text
                    partes = fecha_texto.split("/")
                    mes = str(partes[1])
                    año = str(partes[2])
                    fecha_oficial = False
                except:    
                    print('Exception localizando fecha en la tabla...')
                    print('----------------------------------------------------------------------')
                    reintentos += 1

            #Obtenemos el boton de descarga que haremos click           
            reintentos = 0            
            descarga_oficial = True
            while (descarga_oficial) and reintentos <= 5:
                try:
                    print('Try localizando boton descarga ...', reintentos)
                    print('----------------------------------------------------------------------')
                    reintentos += 1
                    descarga = self.driver.find_element(By.XPATH,f'/html/body/div[1]/div[2]/section/div[2]/div/div/div[2]/div[2]/div/table/tbody/tr[{i}]/td[6]/div')
                    descarga.click()
                    boton_confirmar_descarga = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div[1]/div/div/div/div/button')))
                    boton_confirmar_descarga.click()
                    time.sleep(7)
                    descarga_oficial = False
                except:    
                    print('Exception en el boton descarga...')
                    print('----------------------------------------------------------------------')
                    reintentos += 1
                    
            # Esperar a que se abra la ventana emergente
            intentos = 0
            reintentar_ventana = True
            while (reintentar_ventana):
                try:
                    # Obtiene el identificador de la ventana actual
                    current_window = self.driver.current_window_handle
                    print('Ventana principal: ', current_window)
                    print('----------------------------------------------------------------------')
                    print('Try en la funcion manejo de ventanas abiertas...', intentos)
                    print('----------------------------------------------------------------------')
                    intentos += 1
                    window_handles_all = self.driver.window_handles
                    reintentar_ventana = False
                            
                except:    
                    print('Exception en la funcion manejo de ventanas abiertas')
                    print('----------------------------------------------------------------------')
                    print('----------------------------------------------------------------------')
                    time.sleep(60) #espera para que cargue ventana emergente          
                    reintentar_ventana = intentos <= 3
                                    
            # Obtiene los identificadores de las ventanas abiertas
            self.driver.implicitly_wait(35)
            print('Ventanas abiertas: ', window_handles_all, len(window_handles_all))
            print('----------------------------------------------------------------------')            
                                    
            # Cambiar al manejo de la ventana emergente
            while True:
                for window_handle in window_handles_all:
                    if window_handle != current_window:
                        self.driver.switch_to.window(window_handle)
                        print('Ventana emergente: ', window_handle)
                        print('----------------------------------------------------------------------')
                        break
                break                

            # Esperar hasta que el elemento esté presente en la página
            self.driver.implicitly_wait(35)
            while True:                
                try:
                    # Obtener la URL de la ventana emergente
                    ventana_emergente_url = self.driver.current_url
                    print("URL de la ventana emergente:", ventana_emergente_url)
                    print('----------------------------------------------------------------------')

                    # Realizar una solicitud GET para obtener la data binaria del documento
                    response = requests.get(ventana_emergente_url, stream=True)

                    #Obtengo resultado del diccionario
                    resultado_diccionario = self.diccionario()
                    
                    # Obtener el nombre del archivo a partir de los datos del proceso de descarga
                    folder_path = './input/'
                    file_name = folder_path + str(valor_boleta)+"_"+ resultado_diccionario[mes]+"_"+año+".pdf" 

                    # Guardar la data binaria en un archivo PDF
                    with open(file_name, 'wb') as file:
                        response.raw.decode_content = True
                        shutil.copyfileobj(response.raw, file)
                        print("Guardando archivo:", file_name)
                        print('----------------------------------------------------------------------')
                    break
                                
                except:    
                    print("No se encontró el elemento con el id especificado...")
                    print('----------------------------------------------------------------------')

                    print('Conteo de documentos: ', i)
                    print('----------------------------------------------------------------------')                

            # Cerrar la ventana emergente
            time.sleep(5)
            self.driver.close()
            time.sleep(15)                            

            # Cambiar de nuevo al manejo de ventana principal
            self.driver.switch_to.window(current_window)
            print('Cual ventana es: ', current_window)
            print('----------------------------------------------------------------------')
                        
            time.sleep(10)
                
            print('pasamos al siguiente archivo')
            i +=1

    def archivos(self):
        
        folder_path = './input/'
        output_path = './output/'
        
        #Revisamos si hay archivos pdf en la carpeta input
        archivos_pdf = glob.glob(os.path.join(folder_path, '*.pdf'))

        #Si no encuentra archivos es porque no se realizo la ejecucion correcta y hay que mandar mail
        if not archivos_pdf:
            print(f'No se encontraron archivos PDF en la carpeta "{folder_path}".')
        else:
            #Si encuentra me entregara todos los documentos con los que trabajaremos
            print(f'Se encontraron los siguientes archivos PDF en la carpeta "{folder_path}":')
            
            for archivo in archivos_pdf:
                
                nombre_oficial = archivo.replace('./input','')
                            
                with fitz.open(archivo) as pdf_documento:
                    texto_completo = ''

                    for pagina_num in range(pdf_documento.page_count):
                        pagina = pdf_documento.load_page(pagina_num)
                        texto_completo += pagina.get_text()
                    
                    lista_limpia = [elemento.strip() for elemento in texto_completo.split('\n')]
                    
                    #PRIMERA TABLA
                    #Posicion 1
                    texto_a_verificar = 'N° FACTURA:'
                    posicion_1 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_1 = idx
                            n_factura_bruta = lista_limpia[posicion_1]
                            partes = n_factura_bruta.split(': ')    
                            n_factura = partes[1]      

                    #Posicion 2 y 3
                    elemento_a_buscar = 'Último pago:'
                    try:
                        posicion_2 = lista_limpia.index(elemento_a_buscar)
                        direccion_bruto = lista_limpia[posicion_2+8]
                        partes = direccion_bruto.split()
                        direccion = partes[0] + partes[1] + partes[2]
                        comuna = partes[3]
                    except:
                        print('elemento no se encuentra disponible')
                        direccion = ''
                        comuna = ''
          
                    #Posicion 4
                    texto_a_verificar = 'N° CLIENTE:'
                    posicion_4 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_4 = idx
                            n_cliente_bruta = lista_limpia[posicion_4]
                            partes = n_cliente_bruta.split(': ')    
                            n_cliente = partes[1]  

                    #Posicion 5
                    texto_a_verificar = 'FECHA EMISIÓN:'
                    posicion_5 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_5 = idx
                            fecha_emision_bruto = lista_limpia[posicion_5]
                            partes = fecha_emision_bruto.split(':')
                            fecha_emision_b = partes[1]
                            fecha_emision_ = fecha_emision_b.split()
                            dia = fecha_emision_[0]
                            mes = self.dic_datos(fecha_emision_[1])
                            año = fecha_emision_[2]
                            fecha_emision = dia +'-'+ mes +'-'+ año
                    
                    #SEGUNDA TABLA
                    #Posicion 6
                    texto_a_verificar = 'FECHA VENCIMIENTO:'
                    posicion_6 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_6 = idx
                            fecha_vencimiento_bruto = lista_limpia[posicion_6]
                            partes = fecha_vencimiento_bruto.split(':')
                            fecha_vencimiento_b = partes[1]
                            fecha_vencimiento_ = fecha_vencimiento_b.split()
                            dia = fecha_vencimiento_[0]
                            mes = self.dic_datos(fecha_vencimiento_[1])
                            año = fecha_vencimiento_[2]
                            fecha_vencimiento = dia +'-'+ mes +'-'+ año
                    
                    #Posicion 7
                    elemento_a_buscar = 'Tarifa'
                    try:
                        posicion_7 = lista_limpia.index(elemento_a_buscar)
                        tarifa = lista_limpia[posicion_7+15]
                    except:
                        print('elemento no se encuentra disponible')
                        tarifa = ''
                    
                    #Posicion 8
                    elemento_a_buscar = 'Zona Tarifaria'
                    try:
                        posicion_8 = lista_limpia.index(elemento_a_buscar)
                        zona_tarifaria = lista_limpia[posicion_8+15]
                    except:
                        print('elemento no se encuentra disponible')
                        zona_tarifaria = ''

                    #Posicion 9
                    elemento_a_buscar = 'Subestación'
                    try:
                        posicion_9 = lista_limpia.index(elemento_a_buscar)
                        subestacion = lista_limpia[posicion_9+15]
                    except:
                        print('elemento no se encuentra disponible')
                        subestacion = ''
                        
                    #Posicion 10
                    elemento_a_buscar = 'Potencia Conectada'
                    try:
                        posicion_10 = lista_limpia.index(elemento_a_buscar)
                        potencia_conectada_b = lista_limpia[posicion_10+15]
                        numeros = re.findall(r'\d+', potencia_conectada_b)
                        potencia_conectada = numeros[0]
                    except:
                        print('elemento no se encuentra disponible')
                        potencia_conectada = ''

                    #Posicion 11
                    elemento_a_buscar = 'Dirección'
                    try:
                        posicion_11 = lista_limpia.index(elemento_a_buscar)
                        direccion2 = lista_limpia[posicion_11+14]
                    except:
                        print('elemento no se encuentra disponible')
                        direccion2 = ''
                        
                    #Posicion 12
                    elemento_a_buscar = 'Bloque de Corte'
                    try:
                        posicion_12 = lista_limpia.index(elemento_a_buscar)
                        bloque_corte = lista_limpia[posicion_12+14]
                    except:
                        print('elemento no se encuentra disponible')
                        bloque_corte = ''
                    
                    #Posicion 13
                    elemento_a_buscar = 'Cons. Referencia Prx. Per'
                    try:
                        posicion_13 = lista_limpia.index(elemento_a_buscar)
                        cons_referencia_b = lista_limpia[posicion_13+14]
                        partes = cons_referencia_b.split()
                        cons_referencia =partes[0]
                    except:
                        print('elemento no se encuentra disponible')
                        cons_referencia = ''
                        
                    #Posicion 14
                    elemento_a_buscar = 'Grupo Consumo'
                    try:
                        posicion_14 = lista_limpia.index(elemento_a_buscar)
                        grupo_consumo = lista_limpia[posicion_14+14]
                    except:
                        print('elemento no se encuentra disponible')
                        grupo_consumo = ''
                    
                    #Posicion 15
                    elemento_a_buscar = 'Fecha Límite para cambio tarifa'
                    try:
                        posicion_15 = lista_limpia.index(elemento_a_buscar)
                        fecha_limite_bruto = lista_limpia[posicion_15+14]
                        partes = fecha_limite_bruto.split('-')
                        dia_limite = partes[0]
                        mes_limite = self.dic_datos(partes[1])
                        año_limite = partes[2]
                        fecha_limite_cambi_tar = dia_limite + mes_limite + año_limite
                    except:
                        print('elemento no se encuentra disponible')
                        fecha_limite_cambi_tar = ''                   
                    
                    #Posicion 16
                    elemento_a_buscar = 'Fecha Termino de tarifa'
                    try:
                        posicion_16 = lista_limpia.index(elemento_a_buscar)
                        fecha_termino_bruto = lista_limpia[posicion_16+14]
                        partes = fecha_termino_bruto.split('-')
                        dia_termino = partes[0]
                        mes_termino = self.dic_datos(partes[1])
                        año_termino = partes[2]
                        fecha_termino_cambi_tar = dia_termino + mes_termino + año_termino
                    except:
                        print('elemento no se encuentra disponible')
                        fecha_termino_cambi_tar = ''  

                    #Posicion 17
                    elemento_a_buscar = 'Demanda Leida'
                    try:
                        posicion_17 = lista_limpia.index(elemento_a_buscar)
                        fecha_termino_bruto = lista_limpia[posicion_17+1]
                        partes = fecha_termino_bruto.split('')
                        cantid_demanda_leida = partes[0]
                    except:
                        print('elemento no se encuentra disponible')
                        cantid_demanda_leida = '' 

                    #Posicion 18 y 19
                    elemento_a_buscar = 'LECTURAS'
                    try:
                        posicion_18 = lista_limpia.index(elemento_a_buscar)
                        fecha_lecutras_bruto = lista_limpia[posicion_18+1]
                        fechas = re.findall(r'\d{2} [A-Za-z]{3} \d{4}', fecha_lecutras_bruto)
                        fecha_inicio_bruto = fechas[0]
                        dia_inicio = fecha_inicio_bruto[0]
                        mes_inicio = self.dic_datos(fecha_inicio_bruto[1])
                        año_inicio = fecha_inicio_bruto[2]
                        fecha_inicio = dia_inicio +'-'+ mes_inicio +'-'+ año_inicio
                        fecha_fin_bruto = fechas[1]
                        dia_fin = fecha_fin_bruto[0]
                        mes_fin = self.dic_datos(fecha_fin_bruto[1])
                        año_fin = fecha_fin_bruto[2]
                        fecha_fin = dia_fin +'-'+ mes_fin +'-'+ año_fin
                    except:
                        print('elemento no se encuentra disponible')
                        fecha_inicio = ''
                        fecha_fin = '' 

                    #Posicion 20,21
                    #Debemos hacer una evaluacion para ver el largo
                    elemento_a_buscar = 'Nr. Medidor Cte'
                    posicion_x = lista_limpia.index(elemento_a_buscar)

                    elemento_a_buscar2 = 'Fecha estimada próx. lectura:'
                    posicion_y = lista_limpia.index(elemento_a_buscar2)
                    
                    diferencia = posicion_y-posicion_x
                    
                    if diferencia == 19:
                        try:
                            posicion_20 = lista_limpia.index(elemento_a_buscar)
                            mr = lista_limpia[posicion_20+1]
                            mr_2 = lista_limpia[posicion_20+2]
                            mr_3 = lista_limpia[posicion_20+3]
                            n_medidor_1 = mr+mr_2+mr_3
                            cte_1 = lista_limpia[posicion_20+4]
                        except:
                            print('elemento no se encuentra disponible')
                            n_medidor_1 = ''
                            cte_1 = ''
                    
                    elif diferencia == 14:
                        try:
                            posicion_20 = lista_limpia.index(elemento_a_buscar)
                            n_medidor_1 = lista_limpia[posicion_20+1]
                            cte_1 = lista_limpia[posicion_20+2]
                        except:
                            print('elemento no se encuentra disponible')
                            n_medidor_1 = ''
                            cte_1 = ''

                    ##Posicion 27,28,29,30,31,32 y 33
                    elemento_a_buscar = 'Nr. Medidor Cte'
                    try:
                        posicion_21 = lista_limpia.index(elemento_a_buscar)
                        n_medidor_2 = lista_limpia[posicion_21+8]
                        cte_2 = lista_limpia[posicion_21+9]
                        tipo_2 = lista_limpia[posicion_21+10]
                        prop_2 = lista_limpia[posicion_21+11]
                        actual_2 = lista_limpia[posicion_21+12]
                        anterior_2 = lista_limpia[posicion_21+13]
                        consumo_2 = lista_limpia[posicion_21+14]
                    except:
                        print('elemento no se encuentra disponible')
                        n_medidor_2 = ''
                        cte_2 = ''
                        tipo_2 = ''
                        prop_2 = ''
                        actual_2 = ''
                        anterior_2 = ''
                        consumo_2 = ''

                    #Posicion 34
                    elemento_a_buscar = 'Fecha estimada próx. lectura:'
                    try:
                        posicion_34 = lista_limpia.index(elemento_a_buscar)
                        fecha_estimada_b = lista_limpia[posicion_34+1]
                        partes = fecha_estimada_b.split('-')
                        dia_estimado = partes[0]
                        mes_estimado = self.dic_datos(partes[1])
                        año_estimado = partes[2]
                        fecha_estimada = dia_estimado +'-'+ mes_estimado +'-'+ año_estimado
                    except:
                        print('elemento no se encuentra disponible')
                        fecha_estimada = ''
        
                    #Posicion 35.0
                    elemento_a_buscar = 'Administracion del servicio'
                    try:
                        posicion_35 = lista_limpia.index(elemento_a_buscar)
                        cobro_administracion = lista_limpia[posicion_35+1]
                    except:
                        print('elemento no se encuentra disponible')
                        cobro_administracion = ''
                        
                    #Posicion 35.1
                    elemento_a_buscar = 'Administración del servicio'
                    try:
                        posicion_35 = lista_limpia.index(elemento_a_buscar)
                        cobro_administracion = lista_limpia[posicion_35+1]
                    except:
                        print('elemento no se encuentra disponible')
                        cobro_administracion = ''
                    
                    #Posicion 36 y 39
                    elemento_a_buscar = 'Electricidad consumida'
                    try:
                        posicion_36 = lista_limpia.index(elemento_a_buscar)
                        cantidad_electricidad_con = lista_limpia[posicion_36-1]
                        monto_electricidad_con = lista_limpia[posicion_36+2]
                    except:
                        print('elemento no se encuentra disponible')
                        cantidad_electricidad_con = ''
                        monto_electricidad_con = ''

                    #Posicion 40
                    elemento_a_buscar = 'Uso del Sistema de Transmision'
                    try:
                        posicion_40 = lista_limpia.index(elemento_a_buscar)
                        monto_sistema_transmision = lista_limpia[posicion_40+1]
                    except:
                        print('elemento no se encuentra disponible')
                        monto_sistema_transmision = ''

                    #Posicion 37 y 41
                    elemento_a_buscar = 'Cargo Ley 21.472'
                    try:
                        posicion_37= lista_limpia.index(elemento_a_buscar)
                        cantidad_cargo_ley = lista_limpia[posicion_37-1]
                        monto_cargo_ley = lista_limpia[posicion_37+2]
                    except:
                        print('elemento no se encuentra disponible')
                        cantidad_cargo_ley = ''
                        monto_cargo_ley = ''

                    #Posicion 38
                    elemento_a_buscar = 'Cargo Ley 21.472'
                    try:
                        posicion_37= lista_limpia.index(elemento_a_buscar)
                        cantidad_cargo_ley = lista_limpia[posicion_37-1]
                        monto_cargo_ley = lista_limpia[posicion_37+2]
                    except:
                        print('elemento no se encuentra disponible')
                        cantidad_cargo_ley = ''
                        monto_cargo_ley = ''

                    #Posicion 38 y 42
                    elemento_a_buscar = 'Demanda Máxima'
                    try:
                        posicion_38= lista_limpia.index(elemento_a_buscar)
                        cantidad_demanda_max= lista_limpia[posicion_38-1]
                        monto_demanda_max = lista_limpia[posicion_38+2]
                    except:
                        print('elemento no se encuentra disponible')
                        cantidad_demanda_max = ''
                        monto_demanda_max = ''

                    #Posicion 43
                    elemento_a_buscar = 'Pago Fuera Plazo'
                    try:
                        posicion_43= lista_limpia.index(elemento_a_buscar)
                        pago_fuera_plazo= lista_limpia[posicion_43+1]
                    except:
                        print('elemento no se encuentra disponible')
                        pago_fuera_plazo = ''

                    #Posicion 44
                    elemento_a_buscar = 'Interés por Mora'
                    try:
                        posicion_44= lista_limpia.index(elemento_a_buscar)
                        interes_mora= lista_limpia[posicion_44+1]
                    except:
                        print('elemento no se encuentra disponible')
                        interes_mora = ''
                        
                    #Posicion 45
                    elemento_a_buscar = 'Saldo Anterior (1)'
                    try:
                        posicion_45= lista_limpia.index(elemento_a_buscar)
                        saldo_anterior= lista_limpia[posicion_45+1]
                    except:
                        print('elemento no se encuentra disponible')
                        saldo_anterior = ''
                        
                    #Posicion 46,47,48,49,50,51,52
                    elemento_a_buscar = 'Tarifa'
                    try:
                        posicion_46 = lista_limpia.index(elemento_a_buscar)
                        monto_exento = lista_limpia[posicion_46-1]
                        monto_afecto = lista_limpia[posicion_46-2]
                        iva = lista_limpia[posicion_46-3]
                        total_mes = lista_limpia[posicion_46-4]
                        otros_cargos = lista_limpia[posicion_46-5]
                        total_a_pagar = lista_limpia[posicion_46-6]
                    except:
                        print('elemento no se encuentra disponible')
                        posicion_46 = lista_limpia.index(elemento_a_buscar)
                        monto_exento = ''
                        monto_afecto = ''
                        iva = ''
                        total_mes = ''
                        otros_cargos = ''
                        total_a_pagar = ''                    

                #Cargamos libro excel donde volcaremos los datos
                libro = load_workbook(output_path+'/'+'Formato Planilla.xlsx')
                hoja_electricidad = libro['Electricidad']
                    
                ultima_fila = hoja_electricidad.max_row
                
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=17).value = int(n_factura)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=17).value = n_factura
                    
                hoja_electricidad.cell(row=ultima_fila+1,column=21).value = direccion
                hoja_electricidad.cell(row=ultima_fila+1,column=22).value = comuna
                hoja_electricidad.cell(row=ultima_fila+1,column=8).value = n_cliente
                hoja_electricidad.cell(row=ultima_fila+1,column=23).value = fecha_emision 
                hoja_electricidad.cell(row=ultima_fila+1,column=24).value = fecha_vencimiento
                hoja_electricidad.cell(row=ultima_fila+1,column=10).value = tarifa 
                
                hoja_electricidad.cell(row=ultima_fila+1,column=20).value = subestacion
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=28).value = int(potencia_conectada)
                except:    
                    hoja_electricidad.cell(row=ultima_fila+1,column=28).value = potencia_conectada

                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=36).value = int(cons_referencia)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=36).value = cons_referencia

                hoja_electricidad.cell(row=ultima_fila+1,column=41).value = cantid_demanda_leida
                hoja_electricidad.cell(row=ultima_fila+1,column=14).value = fecha_inicio
                hoja_electricidad.cell(row=ultima_fila+1,column=107).value = n_medidor_1
                
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=32).value = int(cte_1)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=32).value = cte_1

                hoja_electricidad.cell(row=ultima_fila+1,column=27).value = fecha_estimada
                
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=103).value = int(cobro_administracion)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=103).value = cobro_administracion
                
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=33).value = int(cantidad_electricidad_con)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=33).value = cantidad_electricidad_con
                
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=59).value = int(monto_electricidad_con)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=59).value = monto_electricidad_con

                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=97).value = int(monto_sistema_transmision)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=97).value = monto_sistema_transmision

                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=108).value = int(monto_cargo_ley)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=108).value = monto_cargo_ley
                
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=101).value = int(monto_demanda_max)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=101).value = monto_demanda_max   
                
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=92).value = int(pago_fuera_plazo)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=92).value = pago_fuera_plazo
                
                try:   
                    hoja_electricidad.cell(row=ultima_fila+1,column=93).value = int(interes_mora)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=93).value = interes_mora

                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=98).value = int(iva)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=98).value = iva

                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=79).value = int(saldo_anterior)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=79).value = saldo_anterior
                
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=80).value = int(otros_cargos)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=80).value = otros_cargos
                
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=81).value = int(total_a_pagar)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=81).value = total_a_pagar
                
                libro.save(output_path+'/'+'Formato Planilla.xlsx')
                    
                # #Copiamos el archivo a la carpeta outpu con el nombre que corresponde
                # shutil.copy(archivo, output_path+nombre_oficial)
                # print('-------------')
    
        # #Obtenemos los archivos de la carpeta input
        # archivos_en_carpeta = os.listdir(folder_path)

        # # Iterar sobre los archivos y eliminarlos
        # for archivo in archivos_en_carpeta:
        #     ruta_archivo = os.path.join(folder_path, archivo)
        #     if os.path.isfile(ruta_archivo):
        #         os.remove(ruta_archivo)
